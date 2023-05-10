import openpyxl



def open_xls():
    """打开Excel文档"""
    wb = openpyxl.load_workbook('example.xlsx')
    print('wb type: ', type(wb))


# 2.获取工作表
def get_sheet():
    wb = openpyxl.load_workbook('example.xlsx')
    sheet_list = wb.get_sheet_names()  # 获得所有工作表
    print(sheet_list)
    sheet = wb.get_sheet_by_name('Sheet3')
    print('sheet type: ', type(sheet))
    print('sheet name: ', sheet.title)


def get_cell():
    wb = openpyxl.load_workbook('example.xlsx')
    # sheet = wb.get_sheet_by_name('Sheet1')  # 获取指定的表格--该方法已经弃用
    sheet = wb['Sheet1']
    print('cell type: ', type(sheet['A1']))
    # 获取单元格的值
    print("A1 cell:", sheet['A1'].value)
    print("B1 cell:", sheet['B1'].value)
    print("C1 cell:", sheet['C1'].value)
    # 获取单元格的另一种方式
    B1 = sheet.cell(row=1, column=2)
    print(B1.value)
    # 循环使用
    for i in range(1, 8, 2):
        print(i, sheet.cell(row=i, column=2).value)
    # 获取表格的总计行数
    # sheet2 = wb.get_sheet_by_name('Sheet1')
    # row_num = sheet2.
    # print(row_num)
    # column_num = sheet2.get_highest_column()
    # print(column_num)



def get_row_column():
    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb['Sheet1']
    result = tuple(sheet['A1':'C3'])
    print(result)
    for rowOfCellObjects in sheet['A1':'C3']:
        for cellObj in rowOfCellObjects:
            print(cellObj.coordinate, cellObj.value)
        print('-----END OF ROW------')


def get_row_column02():
    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb['Sheet1']
    result = sheet
    print(result)
    print('---------------')
    num_row = sheet.cell(1, 1)
    print(type(num_row))
    for cellObj in sheet[1]:
        print(cellObj.value)


# 获取表格中最大行数
def get_highest_row(sheet):
    # wb = openpyxl.load_workbook('example.xlsx')
    # sheet = wb['Sheet1']
    row_number = 0
    cell = sheet.cell(1, 1)
    print(cell.value)
    if not cell.value:
        return row_number
    else:
        row_number = 1
        row = 2
        while sheet.cell(row, 1).value:
            row_number = row_number + 1
            row = row + 1
    return row_number


# 获取表格中最大列数
def get_highest_col(sheet):
    # wb = openpyxl.load_workbook('example.xlsx')
    # sheet = wb['Sheet1']
    col_number = 0
    cell = sheet.cell(1, 1)
    print(cell.value)
    if not cell.value:
        return col_number
    else:
        col_number = 1
        col = 2
        while sheet.cell(1, col).value:
            col_number = col_number + 1
            col = col + 1
    return col_number






if __name__ == '__main__':
    # open_xls()
    # get_sheet()
    # get_cell()
    # get_row_column02()
    # 测试最大行数
    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb['Sheet1']
    num_row = get_highest_row(sheet)
    print("最大行数: ", num_row)
    num_col = get_highest_col(sheet)
    print('最大列数: ', num_col)



