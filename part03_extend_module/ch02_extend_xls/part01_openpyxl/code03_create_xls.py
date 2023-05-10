# code03 创建表格
import openpyxl


# 1. 创建表格
def test_create_xls():
    wb = openpyxl.Workbook()
    # 获取表格列表
    sheets = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name('Sheet')
    print(sheet.title)
    # 更改表格名称
    sheet.title = 'Spam Bacon Eggs Sheet'
    sheets = wb.get_sheet_names()
    print(sheets)
    # 保存表格
    wb.save('example_copy.xlsx')


def test_create_sheet():
    wb = openpyxl.Workbook()
    sheet_list = wb.get_sheet_names()
    print("表格列表: ", sheet_list)
    result = wb.create_sheet()
    print("type ", result)
    sheet_list = wb.get_sheet_names()
    print("表格列表: ", sheet_list)
    first_sheet = wb.create_sheet(index=0, title='First Sheet')
    sheet_list = wb.get_sheet_names()
    print("表格列表: ", sheet_list)
    wb.create_sheet(index=2, title='Middle Sheet')
    sheet_list = wb.get_sheet_names()
    print("表格列表: ", sheet_list)
    # 移除 表格
    Middle_Sheet = wb.get_sheet_by_name('Middle Sheet')
    wb.remove_sheet(Middle_Sheet)
    sheet_list = wb.get_sheet_names()
    print("表格列表: ", sheet_list)


def write_to_sheet():
    wb = openpyxl.Workbook()
    sheet1 = wb.get_sheet_by_name('Sheet')
    sheet1['A1'] = 'Hello world'
    print(sheet1['A1'].value)


if __name__ == '__main__':
    # test_create_sheet()
    write_to_sheet()
