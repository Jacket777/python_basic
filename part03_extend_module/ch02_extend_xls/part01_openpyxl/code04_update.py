# code04 更新表格数据

import openpyxl
from openpyxl.styles import Font, NamedStyle,PatternFill
from openpyxl.utils import get_column_letter, get_column_interval,column_index_from_string
import Tools as tools

def update_sheet():
    wb = openpyxl.load_workbook('produceSales.xlsx')
    sheet = wb.get_sheet_by_name('Sheet')
    PRICE_UPDATES = {
        'Garlic': 3.07,
        'Celery': 1.19,
        'Lemon': 1.27
    }
    for rowNum in range(2, tools.get_highest_row(sheet)):
        produceName = sheet.cell(row=rowNum, column=1).value
        if produceName in PRICE_UPDATES:
            sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
    wb.save('updateProduceSales.xlsx')


# 新版本无styles 属性
def set_font():
    wb = openpyxl.Workbook()
    sheet = wb.get_sheet_by_name('Sheet')
    italic24_Font = Font(size=24, italic=True)
    style_obj = NamedStyle(font=italic24_Font)
    sheet.cell(row=2,column=8,value="test").font
    sheet['A'].style = style_obj
    sheet['A1'] = 'Hello world!'
    wb.save('styled.xlsx')





if __name__ == '__main__':
    # update_sheet()
    set_font()
